#安排一个作业 #写一个类 类里面有2个方法 1）读数据 2）写数据 
#1）读数据可以读取整个Excel里面所有的数据，每一行数据都放到一个子列表里面，所有子列表数据放到一个大列表里面,要求把读取到的数据返回
#2）写数据可以在Excel里面指定的单元格里面写入指定的值，不需要返回值

#温馨提示：记得关闭和保存Excel



import json
from openpyxl import load_workbook

login_datas=[]
class w_2:
    def __init__(self,excel,sheet):
        self.excel=excel #excel名称
        self.sheet=sheet   #表单名称
    def read(self):
        wb=load_workbook(self.excel) #打开文件
        sheet=wb[self.sheet]  #定位表单
        for i in range (2,sheet.max_row+1):  #获取最大行
            user={
                'url':sheet.cell(i,1).value,
                'method':sheet.cell(i,2).value,
                'data':json.loads(sheet.cell(i,3).value),
                'exp_data':sheet.cell(i,4).value,
                'case_id':sheet.cell(i,7).value
            }
            login_datas.append(user)

        wb.close()
        return login_datas

    def write_result(self,row,result):
        wb=load_workbook(self.excel) #打开文件
        sheet=wb[self.sheet]  #定位表单

        sheet.cell(row, 6, result)

        wb.save(self.excel)
        wb.close()


if __name__=='__main__':
    g=w_2('py15.xlsx','Sheet1')
    a=g.read()
    print(a)
    #
    # g.write_result(3,'haha1')












