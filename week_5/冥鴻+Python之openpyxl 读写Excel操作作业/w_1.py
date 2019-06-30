#安排一个作业 #写一个类 类里面有2个方法 1）读数据 2）写数据 
#1）读数据可以读取整个Excel里面所有的数据，每一行数据都放到一个子列表里面，所有子列表数据放到一个大列表里面,要求把读取到的数据返回
#2）写数据可以在Excel里面指定的单元格里面写入指定的值，不需要返回值

#温馨提示：记得关闭和保存Excel


from openpyxl import load_workbook
q=[]
w=[]
class w_2:
    def __init__(self,excel,bg):
        self.excel=excel #excel名称
        self.bg=bg   #表单名称
    def read(self):
        wb=load_workbook(self.excel) #打开文件
        sheet=wb[self.bg]  #定位表单
        for i in range (1,sheet.max_row+1):  #获取最大行

            for j in range (1,sheet.max_column+1): #获取最大列
                if sheet.cell(i,j).value:
                    c=sheet.cell(i,j).value
                    q.append(c)
                    # print(q)
            w.append(list(q))
            q.clear()
        return w
    def write(self,row,column,neirong):
        # self.row = row
        # self.column = column
        # self.neirong = neirong
        wb = load_workbook(self.excel)  # 打开文件
        sheet = wb[self.bg]  # 定位表单
        sheet.cell(row,column,neirong)
        wb.save(self.excel)
        wb.close()


if __name__=='__main__':
    g=w_2('python15.xlsx','Sheet1')
    a=g.read()
    print(a)
    g.write(2,3,'买过去')











